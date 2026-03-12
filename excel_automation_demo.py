#!/usr/bin/env python3
"""
Excel Automation Demo - Stock Portfolio Tracker
This script creates an automated Excel file that tracks stock portfolio
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule
from openpyxl.utils import get_column_letter
import datetime
import os

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Portfolio"

# Styles
header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=12)
green_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
red_fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Headers
headers = ["股票代码", "股票名称", "买入日期", "买入价", "当前价", "数量", "买入金额", "当前价值", "盈亏", "盈亏%"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# Sample Data (demo portfolio)
sample_data = [
    ["600036", "招商银行", "2026-01-15", 38.50, 39.23, 100, 3850, 3923, 73, 1.89],
    ["600030", "中信证券", "2026-01-20", 26.80, 26.90, 200, 5360, 5380, 20, 0.37],
    ["000001", "平安银行", "2026-02-01", 12.50, 12.65, 300, 3750, 3795, 45, 1.20],
    ["600009", "上海机场", "2026-02-10", 30.20, 29.93, 150, 4530, 4489.5, -40.5, -0.89],
    ["601012", "隆基绿能", "2026-02-15", 18.00, 17.76, 200, 3600, 3552, -48, -1.33],
]

# Fill data
for row_idx, row_data in enumerate(sample_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        
        # Format numbers
        if col_idx in [4, 5, 7, 8]:  # Price columns
            cell.number_format = '¥#,##0.00'
        elif col_idx == 9:  # P&L amount
            cell.number_format = '¥#,##0.00'
        elif col_idx == 10:  # P&L %
            cell.number_format = '0.00%'
    
    # Conditional formatting for P&L
    pl_cell = ws.cell(row=row_idx, column=9)
    pl_pct_cell = ws.cell(row=row_idx, column=10)
    
    if pl_cell.value and pl_cell.value > 0:
        pl_cell.fill = green_fill
        pl_cell.font = Font(color="FFFFFF", bold=True)
    elif pl_cell.value and pl_cell.value < 0:
        pl_cell.fill = red_fill
        pl_cell.font = Font(color="FFFFFF", bold=True)

# Summary Section
summary_start = len(sample_data) + 3
ws.cell(row=summary_start, column=1, value="📊 投资汇总").font = Font(bold=True, size=14)

summary_labels = ["总投入", "当前总值", "总盈亏", "盈亏率"]
summary_formulas = [
    f"=SUM(G2:G{len(sample_data)+1})",
    f"=SUM(H2:H{{}})".format(len(sample_data)+1),
    f"=SUM(I2:I{len(sample_data)+1})",
    f"=I{summary_start+4}/G{summary_start+2}"
]

for i, (label, formula) in enumerate(zip(summary_labels, summary_formulas), 1):
    ws.cell(row=summary_start + i, column=1, value=label).font = Font(bold=True)
    cell = ws.cell(row=summary_start + i, column=2, value=formula.replace("{{}}", str(len(sample_data)+1)))
    if i == 4:
        cell.number_format = '0.00%'
    else:
        cell.number_format = '¥#,##0.00'
    cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

# Adjust column widths
column_widths = [12, 15, 12, 12, 12, 10, 15, 15, 12, 12]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Add a chart
chart = BarChart()
chart.title = "股票盈亏对比"
chart.y_axis.title = "盈亏金额"
chart.x_axis.title = "股票"

data = Reference(ws, min_col=9, min_row=1, max_row=len(sample_data)+1)
cats = Reference(ws, min_col=2, min_row=2, max_row=len(sample_data)+1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.width = 15
chart.height = 10
ws.add_chart(chart, "L2")

# Create second sheet for signals
ws2 = wb.create_sheet("交易信号")
ws2["A1"] = "📈 今日交易信号"
ws2["A1"].font = Font(bold=True, size=16)

signal_headers = ["信号", "股票", "价格", "原因", "状态"]
for col, header in enumerate(signal_headers, 1):
    cell = ws2.cell(row=2, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font

# Sample signals
signals = [
    ["🔴 卖出", "600030 中信证券", "¥26.90", "触及止盈点+10%", "待执行"],
    ["🟢 买入", "000001 平安银行", "¥12.65", "站上20日均线", "观察中"],
    ["🟡 持有", "600036 招商银行", "¥39.23", "正常上涨趋势", "持有"],
]

for row_idx, signal in enumerate(signals, 3):
    for col_idx, value in enumerate(signal, 1):
        ws2.cell(row=row_idx, column=col_idx, value=value)

# Adjust widths
for col in range(1, 6):
    ws2.column_dimensions[get_column_letter(col)].width = 18

# Save file
output_path = os.path.expanduser("~/Desktop/Stock_Analysis/demo_portfolio.xlsx")
wb.save(output_path)

print(f"✅ Excel自动化示例已创建: {output_path}")
print(f"📊 包含内容:")
print(f"   - 股票投资组合追踪表")
print(f"   - 自动计算盈亏")
print(f"   - 条件格式(绿色=盈利, 红色=亏损)")
print(f"   - 汇总统计")
print(f"   - 可视化图表")
print(f"   - 交易信号表")
print(f"\n🎯 可作为演示项目展示给客户!")
