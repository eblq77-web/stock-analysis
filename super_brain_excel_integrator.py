#!/usr/bin/env python3
"""
Super Brain Pro V3 + Excel Integration
Reads Super Brain analysis and generates Excel reports
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
import os
import glob
from datetime import datetime

# Configuration
DAILY_OVERVIEW_PATH = "~/Desktop/Stock_Analysis/daily_overview/"
OUTPUT_PATH = "~/Desktop/Stock_Analysis/super_brain_reports/"

# Styles
header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=12)
green_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
yellow_fill = PatternFill(start_color="FBBF24", end_color="FBBF24", fill_type="solid")
red_fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_super_brain_excel():
    """Create Excel from Super Brain analysis"""
    
    # Ensure output directory exists
    os.makedirs(os.path.expanduser(OUTPUT_PATH), exist_ok=True)
    
    wb = Workbook()
    
    # Sheet 1: Market Intelligence
    ws1 = wb.active
    ws1.title = "市场情报"
    
    # Headers
    ws1["A1"] = "🧠 Super Brain Pro V3 市场分析"
    ws1["A1"].font = Font(bold=True, size=16)
    ws1.merge_cells("A1:F1")
    
    market_headers = ["指标", "数值", "信号强度", "建议", "更新时间"]
    for col, header in enumerate(market_headers, 1):
        cell = ws1.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    # Sample market data (in real use, parse from Super Brain output)
    market_data = [
        ["市场情绪", "乐观", "🟢 强", "持股待涨", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["趋势方向", "上涨", "🟢 强", "逢低买入", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["波动率", "中等", "🟡 中", "谨慎操作", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["资金流向", "净流入", "🟢 强", "积极做多", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["风险等级", "中等", "🟡 中", "控制仓位", datetime.now().strftime("%Y-%m-%d %H:%M")],
    ]
    
    for row_idx, row_data in enumerate(market_data, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            
            # Color code based on signal
            if "🟢" in str(value):
                cell.fill = green_fill
                cell.font = Font(color="FFFFFF", bold=True)
            elif "🟡" in str(value):
                cell.fill = yellow_fill
    
    # Adjust column widths
    for col in range(1, 6):
        ws1.column_dimensions[get_column_letter(col)].width = 18
    
    # Sheet 2: Stock Picks
    ws2 = wb.create_sheet("股票推荐")
    
    ws2["A1"] = "📈 Super Brain 精选股票"
    ws2["A1"].font = Font(bold=True, size=16)
    ws2.merge_cells("A1:H1")
    
    pick_headers = ["排名", "代码", "名称", "行业", "得分", "信号", "目标价", "止损价"]
    for col, header in enumerate(pick_headers, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    # Sample stock picks (integrate with Super Brain scanner)
    stock_picks = [
        [1, "600036", "招商银行", "金融", 92, "🟢 买入", 42.00, 35.00],
        [2, "300750", "宁德时代", "新能源", 89, "🟢 买入", 220, 180],
        [3, "000651", "格力电器", "家电", 87, "🟡 持有", 40, 34],
        [4, "002475", "立讯精密", "科技", 85, "🟢 买入", 38, 28],
        [5, "601012", "隆基绿能", "新能源", 82, "🟡 持有", 20, 15],
        [6, "300122", "智飞生物", "医药", 80, "🟡 观望", 45, 35],
        [7, "600009", "上海机场", "交运", 78, "🔴 卖出", 32, 28],
        [8, "000001", "平安银行", "金融", 75, "🟡 持有", 14, 11],
    ]
    
    for row_idx, row_data in enumerate(stock_picks, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            
            # Score coloring
            if col_idx == 5:  # Score column
                if value >= 85:
                    cell.fill = green_fill
                    cell.font = Font(color="FFFFFF", bold=True)
                elif value >= 75:
                    cell.fill = yellow_fill
            
            # Signal coloring
            if col_idx == 6:  # Signal column
                if "🟢" in str(value):
                    cell.fill = green_fill
                elif "🔴" in str(value):
                    cell.fill = red_fill
    
    for col in range(1, 9):
        ws2.column_dimensions[get_column_letter(col)].width = 14
    
    # Sheet 3: Portfolio Tracking
    ws3 = wb.create_sheet("持仓追踪")
    
    ws3["A1"] = "💰 持仓组合追踪"
    ws3["A1"].font = Font(bold=True, size=16)
    ws3.merge_cells("A1:J1")
    
    portfolio_headers = ["股票", "代码", "买入价", "当前价", "数量", "市值", "盈亏", "盈亏%", "持仓天数", "状态"]
    for col, header in enumerate(portfolio_headers, 1):
        cell = ws3.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    # Sample portfolio
    portfolio = [
        ["招商银行", "600036", 38.50, 39.23, 100, 3923, 73, 1.89, 15, "🟢 盈利"],
        ["中信证券", "600030", 27.16, 26.90, 200, 5380, -52, -0.96, 10, "🔴 亏损"],
        ["平安银行", "000001", 12.50, 12.65, 300, 3795, 45, 1.20, 20, "🟢 盈利"],
        ["上海机场", "600009", 30.05, 29.93, 150, 4489.5, -18, -0.40, 5, "🟡 持平"],
        ["隆基绿能", "601012", 18.00, 17.76, 200, 3552, -48, -1.33, 3, "🔴 亏损"],
    ]
    
    for row_idx, row_data in enumerate(portfolio, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            
            # Format numbers
            if col_idx in [3, 4, 6]:  # Price columns
                cell.number_format = '¥#,##0.00'
            elif col_idx == 7:  # P&L amount
                cell.number_format = '¥#,##0.00'
            elif col_idx == 8:  # P&L %
                cell.number_format = '0.00%'
            
            # P&L coloring
            if col_idx == 7 and isinstance(value, (int, float)):
                if value > 0:
                    cell.fill = green_fill
                    cell.font = Font(color="FFFFFF")
                elif value < 0:
                    cell.fill = red_fill
                    cell.font = Font(color="FFFFFF")
    
    for col in range(1, 11):
        ws3.column_dimensions[get_column_letter(col)].width = 12
    
    # Sheet 4: Sector Analysis
    ws4 = wb.create_sheet("板块分析")
    
    ws4["A1"] = "📊 板块资金流向"
    ws4["A1"].font = Font(bold=True, size=16)
    ws4.merge_cells("A1:E1")
    
    sector_headers = ["板块", "资金流入", "涨幅", "主力净流入", "建议"]
    for col, header in enumerate(sector_headers, 1):
        cell = ws4.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    sector_data = [
        ["新能源", "12.5亿", "+3.2%", "8.2亿", "🟢 重点关注"],
        ["半导体", "8.3亿", "+2.1%", "5.1亿", "🟢 关注"],
        ["医药", "5.2亿", "+1.5%", "2.8亿", "🟡 观望"],
        ["金融", "-3.1亿", "-0.8%", "-5.2亿", "🔴 回避"],
        ["消费", "2.1亿", "+0.5%", "0.8亿", "🟡 持有"],
    ]
    
    for row_idx, row_data in enumerate(sector_data, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            
            # Color code based on recommendation
            if "🟢" in str(value):
                cell.fill = green_fill
                cell.font = Font(color="FFFFFF")
            elif "🔴" in str(value):
                cell.fill = red_fill
                cell.font = Font(color="FFFFFF")
    
    for col in range(1, 6):
        ws4.column_dimensions[get_column_letter(col)].width = 16
    
    # Sheet 5: Signals Summary
    ws5 = wb.create_sheet("信号汇总")
    
    ws5["A1"] = "🎯 今日交易信号"
    ws5["A1"].font = Font(bold=True, size=16)
    ws5.merge_cells("A1:F1")
    
    signal_headers = ["时间", "股票", "信号", "价格", "原因", "状态"]
    for col, header in enumerate(signal_headers, 1):
        cell = ws5.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    signals = [
        ["09:31", "600036 招商银行", "🟢 买入", 38.72, "站上均线", "已执行"],
        ["09:45", "300750 宁德时代", "🟢 买入", 215.5, "量价齐升", "已执行"],
        ["10:15", "601012 隆基绿能", "🔴 卖出", 18.20, "触及止损", "已执行"],
        ["11:30", "000651 格力电器", "🟡 持有", 37.50, "正常波动", "观察中"],
    ]
    
    for row_idx, row_data in enumerate(signals, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws5.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            
            if col_idx == 3:
                if "🟢" in str(value):
                    cell.fill = green_fill
                    cell.font = Font(color="FFFFFF")
                elif "🔴" in str(value):
                    cell.fill = red_fill
    
    for col in range(1, 7):
        ws5.column_dimensions[get_column_letter(col)].width = 16
    
    # Save file
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    filename = f"SuperBrain_Report_{timestamp}.xlsx"
    output_file = os.path.expanduser(OUTPUT_PATH + filename)
    wb.save(output_file)
    
    print(f"✅ Super Brain Excel报告已生成!")
    print(f"📁 文件位置: {output_file}")
    print(f"\n📊 包含工作表:")
    print(f"   1. 市场情报 - Market intelligence")
    print(f"   2. 股票推荐 - Stock picks from Super Brain")
    print(f"   3. 持仓追踪 - Portfolio tracking")
    print(f"   4. 板块分析 - Sector analysis")
    print(f"   5. 信号汇总 - Today's trading signals")
    
    return output_file

if __name__ == "__main__":
    create_super_brain_excel()
