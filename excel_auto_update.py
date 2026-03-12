#!/usr/bin/env python3
"""
Excel自动更新脚本 - 每日自动更新股价和信号
配合cron job可以每日自动运行
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import datetime
import time
import subprocess

# 读取Excel
def update_portfolio():
    print(f"🔄 [{datetime.datetime.now().strftime('%H:%M:%S')}] 正在更新股票数据...")
    
    try:
        wb = openpyxl.load_workbook('~/Desktop/Stock_Analysis/demo_portfolio.xlsx')
        ws = wb['Portfolio']
        
        # 获取当前股价 (这里用模拟数据 - 实际可以接入API)
        stock_prices = {
            "600036": 39.23,
            "600030": 26.90,
            "000001": 12.65,
            "600009": 29.93,
            "601012": 17.76,
        }
        
        # 更新股价
        for row in range(2, ws.max_row + 1):
            code = ws.cell(row=row, column=1).value
            if code in stock_prices:
                # 更新当前价
                ws.cell(row=row, column=5).value = stock_prices[code]
                ws.cell(row=row, column=5).number_format = '¥#,##0.00'
                
                # 计算当前价值
                quantity = ws.cell(row=row, column=6).value
                current_value = stock_prices[code] * quantity
                ws.cell(row=row, column=8).value = current_value
                ws.cell(row=row, column=8).number_format = '¥#,##0.00'
                
                # 计算盈亏
                buy_price = ws.cell(row=row, column=4).value
                buy_amount = ws.cell(row=row, column=7).value
                pl = current_value - buy_amount
                ws.cell(row=row, column=9).value = pl
                ws.cell(row=row, column=9).number_format = '¥#,##0.00'
                
                # 盈亏百分比
                pl_pct = pl / buy_amount if buy_amount else 0
                ws.cell(row=row, column=10).value = pl_pct / 100
                ws.cell(row=row, column=10).number_format = '0.00%'
                
                # 更新颜色
                if pl > 0:
                    ws.cell(row=row, column=9).fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
                elif pl < 0:
                    ws.cell(row=row, column=9).fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
        
        # 更新汇总
        summary_row = ws.max_row + 2
        
        # 保存
        wb.save('~/Desktop/Stock_Analysis/demo_portfolio.xlsx')
        print(f"✅ 股价更新完成!")
        
    except Exception as e:
        print(f"❌ 更新失败: {e}")

# 生成报告
def generate_report():
    print(f"📊 [{datetime.datetime.now().strftime('%H:%M:%S')}] 正在生成报告...")
    
    report = f"""
================================================================================
                    股票投资组合日报 - {datetime.datetime.now().strftime('%Y-%m-%d')}
================================================================================

📈 持仓概况:
"""
    
    try:
        wb = openpyxl.load_workbook('~/Desktop/Stock_Analysis/demo_portfolio.xlsx')
        ws = wb['Portfolio']
        
        total_invested = 0
        total_value = 0
        
        for row in range(2, ws.max_row + 1):
            code = ws.cell(row=row, column=1).value
            name = ws.cell(row=row, column=2).value
            buy_price = ws.cell(row=row, column=4).value
            current_price = ws.cell(row=row, column=5).value
            quantity = ws.cell(row=row, column=6).value
            pl = ws.cell(row=row, column=9).value
            pl_pct = ws.cell(row=row, column=10).value
            
            if code:
                invested = buy_price * quantity
                value = current_price * quantity
                total_invested += invested
                total_value += value
                
                emoji = "🟢" if pl >= 0 else "🔴"
                report += f"  {emoji} {code} {name}: 买入{buy_price} → 当前{current_price} | 数量:{quantity} | 盈亏:{pl:.2f} ({pl_pct:.2%})\n"
        
        total_pl = total_value - total_invested
        total_pl_pct = (total_pl / total_invested * 100) if total_invested > 0 else 0
        
        report += f"""
--------------------------------------------------------------------------------
💰 汇总:
   总投入: ¥{total_invested:,.2f}
   当前价值: ¥{total_value:,.2f}
   总盈亏: ¥{total_pl:,.2f} ({total_pl_pct:.2f}%)
================================================================================
"""
        
        # 保存报告
        with open('~/Desktop/Stock_Analysis/daily_report.txt', 'w', encoding='utf-8') as f:
            f.write(report)
        
        print(report)
        print(f"✅ 报告已保存到 daily_report.txt")
        
    except Exception as e:
        print(f"❌ 报告生成失败: {e}")

if __name__ == "__main__":
    print("🚀 Excel自动化脚本启动!")
    print("=" * 50)
    
    # 更新数据
    update_portfolio()
    
    # 生成报告
    generate_report()
    
    print("=" * 50)
    print("🎉 完成!")
